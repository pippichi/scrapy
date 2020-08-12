# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

import datetime
import os
import openpyxl
from deya_cofeed.tools import gen_excel
import configparser as cp


class DeyaCofeedPipeline(object):

    def __init__(self):
        self.excel = openpyxl.Workbook()
        self.excel.remove(self.excel['Sheet'])
        self.today = datetime.date.today()

    def process_item(self, item, spider):
        try:
            if spider.name == 'deya_cofeed_s':
                content = item['content']
                gen_excel(content, self.excel)
                return item
            if spider.name == 'deya_cofeed_history_s' or spider.name == 'deya_cofeed_history2_s':
                file_name = item['date'] + '天下粮仓.xlsx'
                if os.path.exists(file_name):
                    excel = openpyxl.load_workbook(file_name)
                else:
                    excel = openpyxl.Workbook()
                    excel.remove(excel['Sheet'])

                gen_excel(item['content'], excel)
                excel.save(file_name)
                return item
        except Exception as e:
            print('!' * 30)
            print(e)
            print('!' * 30)

    def close_spider(self, spider):
        if spider.name == 'deya_cofeed_s':
            try:
                cf = cp.ConfigParser()
                cf.read(os.getcwd() + "/deya_cofeed/config.ini")
                file_path = cf.get('file_location', 'FILE_PATH')
                self.excel.save(file_path + str(self.today) + "天下粮仓.xlsx")
            except Exception as e:
                print('!' * 30)
                print(e)
                print('!' * 30)
        if spider.name == 'deya_cofeed_history_s' or spider.name == 'deya_cofeed_history2_s':
            pass
