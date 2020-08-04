# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

import datetime
import openpyxl


class DeyaCofeedPipeline(object):

    def __init__(self):
        self.excel = openpyxl.Workbook()
        self.excel.remove(self.excel['Sheet'])
        self.today = datetime.date.today()

    def process_item(self, item, spider):
        content = item['content']
        first_row_name = None
        if len(content[0]) == 5:
            sheet = self.excel.create_sheet("国内生猪及仔猪市场交易日报")
            for i in range(len(content)-1):
                for j in range(len(content[i])):
                    if len(content[i]) == 5:
                        sheet.cell(i+1, j+1).value = content[i][j]
                        first_row_name = content[i][0]
                    else:
                        sheet.cell(i+1, 1).value = first_row_name
                        sheet.cell(i+1, j+2).value = content[i][j]
            for j in range(len(content[len(content)-1])):
                if j == 1 or j == 2:
                    continue
                else:
                    sheet.cell(len(content), j+1).value = content[len(content)-1][j]
        if len(content[0]) == 7:
            sheet = self.excel.create_sheet("国内肉鸡市场交易日报")
            for i in range(len(content) - 1):
                for j in range(len(content[i])):
                    if len(content[i]) == 7:
                        sheet.cell(i + 1, j + 1).value = content[i][j]
                        first_row_name = content[i][0]
                    else:
                        sheet.cell(i + 1, 1).value = first_row_name
                        sheet.cell(i + 1, j + 2).value = content[i][j]
            for j in range(len(content[len(content) - 1])):
                sheet.cell(len(content), j + 2).value = content[len(content) - 1][j]
        return item

    def close_spider(self, spider):
        self.excel.save(str(self.today) + "天下粮仓.xlsx")
