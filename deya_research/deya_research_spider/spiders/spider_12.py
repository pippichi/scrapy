import datetime
import random
import time
from decimal import Decimal

import parsel
from openpyxl import Workbook
import pandas as pd
import scrapy
from scrapy import FormRequest
from deya_research_spider.tools import code_verify, insert_value, parse_date
import requests


class Spider12Spider(scrapy.Spider):
    """
    12.全国PP装置生产情况汇总
    """
    name = 'spider_12'
    allowed_domains = ['oilchem.net']
    search_url = 'https://list.oilchem.net/329/559/'
    login_succeed_url = 'https://passport.oilchem.net/member/login/login'
    timestamp = int(round(time.time() * 1000))
    img_url = 'https://passport.oilchem.net/member/login/getImgCode?timestamp=' + str(timestamp)
    code_verify_url = 'https://passport.oilchem.net/member/login/checkImgCode?code={code}'

    username = 'rockyeah'
    password = 'cc7da6bca8aa4a5d9c0ebea54fb566ae'
    errorPaw = 'deya1589'

    # 指标在数据库中的id
    capacity_of_production_id = 30
    repair_id = 32
    capacity_sub_repair_id = 31

    # Sheet name
    sheet_name = "全国PP装置生产情况汇总"

    # 模式1：爬历史数据；模式0：爬最新数据
    mode = 0

    @property
    def start_requests(self):
        try:
            res = parsel.Selector(requests.post("https://search.oilchem.net/solrSearch/select.htm", headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.125 Safari/537.36'},
                                                data={'pageNo': 1, 'keyword': '全国PP装置生产情况汇总'}).text)

            if self.mode == 1:
                urls = []
                last_page = res.xpath('//*[@id="simpledatatable_info"]/text()').extract_first()[6: 8]
                for i in range(1, int(last_page) + 1):
                    res = parsel.Selector(requests.post("https://search.oilchem.net/solrSearch/select.htm", headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.125 Safari/537.36'}, data={'pageNo': i, 'keyword': '全国PP装置生产情况汇总'}).text)
                    temp_urls = list(set(res.xpath('/html/body/div/div[3]/div[1]/div[2]/ul/li//a/@href').extract()))
                    time.sleep(5 + random.uniform(1, 10))
                    for tu in temp_urls:
                        urls.append(tu)
                print("共爬取 " + str(len(urls)) + ' 条url')
                for target in urls:
                    # 识别错误次数
                    count = 0
                    if target[0: 4] != 'http':
                        target = 'https:' + target
                    temp = code_verify(self.img_url, self.code_verify_url)
                    while temp.text != 'true':
                        count += 1
                        print("第{}次识别出错。".format(count))
                        temp = code_verify(self.img_url, self.code_verify_url)
                    yield FormRequest(url=self.login_succeed_url, formdata={'username': self.username, 'password': self.password, 'target': target, 'errorPaw': self.errorPaw}, callback=self.parse)
                    time.sleep(10 + int(random.uniform(1, 10)))
            else:
                url = res.xpath('/html/body/div/div[3]/div[1]/div[2]/ul/li[1]/h2/a/@href').extract_first()
                date = res.xpath('/html/body/div/div[3]/div[1]/div[2]/ul/li[1]/div/div/span/span/text()').extract_first()
                today = datetime.date.today().strftime('%Y-%m-%d')
                if today != date:
                    print("当日全国PP数据还没有出来")
                    return
                else:
                    count = 0
                    if url[0: 4] != 'http':
                        url = 'https:' + url
                    temp = code_verify(self.img_url, self.code_verify_url)
                    while temp.text != 'true':
                        count += 1
                        print("第{}次识别出错。".format(count))
                        temp = code_verify(self.img_url, self.code_verify_url)
                    yield FormRequest(url=self.login_succeed_url,
                                      formdata={'username': self.username, 'password': self.password, 'target': url,
                                                'errorPaw': self.errorPaw}, callback=self.parse)

        except Exception as e:
            print('!' * 30)
            print('step1')
            print(e)
            print('!' * 30)

    def parse(self, response, **kwargs):
        try:
            title = response.xpath('//*[@id="content"]/table/tbody/tr[@class="firstRow"]/td//text()').getall()
            if not '用途' in title:
                pass
            else:
                # 经过观察，表的第一行里面要么有“产能”，要么有“产能（万吨/年）”
                flag = 0 if '产能' in title else 1
                date = response.xpath("/html/body/div[8]/div[2]/div[1]/div[1]/div[1]/span/text()").get()
                date = parse_date(date)
                temp = response.xpath("//*[@id='content']/table/tbody/tr[not(contains(@class, 'firstRow'))]")
                content = []
                for t in temp:
                    td = t.xpath('.//td')
                    temp_list = []
                    for t_td in td:
                        text = t_td.xpath(".//text()").getall()
                        temp_str = "".join(text)
                        temp_list.append(temp_str)
                    content.append(temp_list)
                wb = Workbook()
                wb.remove(wb['Sheet'])
                st = wb.create_sheet(self.sheet_name)
                row = 2
                last_name_col1, last_name_col2, last_name_col3 = None, None, None

                for ti in range(len(title)):
                    st.cell(1, ti + 1).value = title[ti]

                for c in content:
                    if len(c) == len(title):
                        last_name_col1 = c[0]
                        last_name_col2 = c[1]
                        last_name_col3 = c[2]
                        for col in range(len(c)):
                            st.cell(row, col + 1).value = c[col]
                    elif len(c) == len(title) - 1 and len(title) - 1 > 0:
                        st.cell(row, 1).value = last_name_col1
                        last_name_col2 = c[0]
                        last_name_col3 = c[1]
                        for col in range(len(c)):
                            st.cell(row, col + 2).value = c[col]
                    elif len(c) == len(title) - 2 and len(title) - 2 > 0:
                        st.cell(row, 1).value = last_name_col1
                        st.cell(row, 2).value = last_name_col2
                        last_name_col3 = c[0]
                        for col in range(len(c)):
                            st.cell(row, col + 3).value = c[col]
                    else:
                        st.cell(row, 1).value = last_name_col1
                        st.cell(row, 2).value = last_name_col2
                        st.cell(row, 3).value = last_name_col3
                        for col in range(len(c)):
                            st.cell(row, col + 4).value = c[col]
                    row += 1
                wb.save(date + '.xlsx')

                ex_pd = pd.read_excel(date + '.xlsx', sheet_name=self.sheet_name)
                if flag == 0:
                    ex_pd_c = ex_pd['产能']
                    ex_pd_c_filter = ex_pd[ex_pd['用途'] == '--']['产能']
                else:
                    ex_pd_c = ex_pd['产能（万吨/年）']
                    ex_pd_c_filter = ex_pd[ex_pd['用途'] == '--']['产能（万吨/年）']
                capacity_of_production = Decimal(sum(ex_pd_c)).quantize(Decimal("0.0000"))
                repair = Decimal(sum(ex_pd_c_filter) / 365).quantize(Decimal("0.0000"))
                capacity_sub_repair = Decimal(capacity_of_production / 365 - repair).quantize(Decimal("0.0000"))

                insert_value(date, capacity_of_production, self.capacity_of_production_id)
                insert_value(date, capacity_sub_repair, self.capacity_sub_repair_id)
                insert_value(date, repair, self.repair_id)
        except Exception as e:
            print('!' * 30)
            print('step2')
            print(e)
            print('!' * 30)
