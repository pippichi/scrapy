# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import datetime
import pandas as pd

from openpyxl import Workbook
import openpyxl

from deya_PP.properties import PP_EXCEL, PP_EXCEL_TEMPLATE, PP_AIM_FILE
from deya_PP.tools import parse_en, succeed_and_send_email


class DeyaPpPipeline(object):

    def __init__(self):
        self.today = datetime.date.today()
        self.middle_file_name = 'PP_' + str(self.today) + '.xlsx'
        self.origin_excel = openpyxl.load_workbook(PP_EXCEL_TEMPLATE)
        self.origin_excel_sheet = self.origin_excel['PP排产']
        self.origin_excel_sheet.cell(4, 1).value = self.today.strftime("%Y/%m/%d")
        self.MAX_COL = self.origin_excel_sheet.max_column

    def process_item(self, item, spider):
        wb = Workbook()
        ws = wb['Sheet']
        ws.title = "PP粒"
        title = item['title']
        content = item['content']
        ws.append(t for t in title)
        row_index = 2
        for c in content:
            if len(c) == 5:
                for i in range(1, len(c)+1):
                    ws.cell(row_index, i).value = c[i-1]
            elif len(c) == 4:
                for i in range(1, len(c)+1):
                    ws.cell(row_index, i+1).value = c[i-1]
            else:
                for i in range(1, len(c)+1):
                    ws.cell(row_index, i+2).value = c[i-1]
            row_index += 1
        wb.save(self.middle_file_name)
        df = pd.read_excel(self.middle_file_name, sheet_name=ws.title, index_col=False)
        part1 = df.iloc[:, 1]
        insert1 = []
        temp_name = None
        temp_index = 0
        for p in part1:
            if pd.isnull(p):
                insert1 = insert1[:-1]
                insert1.append(temp_name + str(temp_index) + '#')
                temp_index += 1
                insert1.append(temp_name + str(temp_index) + '#')
            else:
                temp_name = p
                temp_index = 1
                insert1.append(p)
        df.insert(2, '石化名称备注', insert1)
        temp_file = pd.read_excel(PP_EXCEL, sheet_name='PP排号', index_col=False)
        num = [n for n in temp_file['型号']]
        type_name = [t for t in temp_file['类别']]
        num_type = dict(zip(num, type_name))
        equip_dyn = parse_en(insert1, df['装置动态'])
        enterprise_type = {}
        for key in equip_dyn:
            flag = 0
            for t in equip_dyn[key]:
                if t.startswith("停车"):
                    enterprise_type[key] = "停车"
                    flag = -1
                    continue
                type_name = num_type.get(t, None)
                if type_name:
                    flag = 1
                    enterprise_type[key] = type_name
                    continue
            if flag == 0:
                enterprise_type[key] = "找不到对应类型"
        for i in range(2, self.MAX_COL):
            self.origin_excel_sheet.cell(4, i).value = enterprise_type.get(self.origin_excel_sheet.cell(1, i).value)

        return item

    def close_spider(self, spider):
        self.origin_excel.save(PP_AIM_FILE)
        succeed_and_send_email(PP_AIM_FILE)
