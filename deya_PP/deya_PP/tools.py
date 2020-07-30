import datetime
import os
import re
import smtplib

import openpyxl
import parsel
import pytesseract
import requests
import pandas as pd

from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from urllib.request import urlretrieve
from PIL import ImageDraw, Image
from openpyxl import Workbook
from deya_PP.settings import EMAIL_HOST_USER, EMAIL_HOST, EMAIL_PORT, EMAIL_HOST_PASSWORD, EMAIL_TO, \
    PP_EXCEL_TEMPLATE, PE_EXCEL_TEMPLATE, PP_AIM_FILE


# 解析爬取的内容
def parse_en(enterprise, content):
    pattern = re.compile(r"[-\dA-Za-z]+", re.S)
    res = []
    for c in content:
        temp = re.findall(pattern, c)
        if len(temp) == 0 or "停车" in c:
            res.append(["停车:" + c])
        else:
            res.append(temp)
    return dict(zip(enterprise, res))


# 二值处理
def get_table(threshold=115):
    table = []
    for i in range(256):
        if i < threshold:
            table.append(0)
        else:
            table.append(1)
    return table


# 图片降噪
def get_pixel(image, x, y, G, N):

    L = image.getpixel((x, y))

    if L > G:
        L = True
    else:
        L = False

    near_dots = 0

    if L == (image.getpixel((x - 1, y - 1)) > G):
        near_dots += 1
    if L == (image.getpixel((x - 1, y)) > G):
        near_dots += 1
    if L == (image.getpixel((x - 1, y + 1)) > G):
        near_dots += 1
    if L == (image.getpixel((x, y - 1)) > G):
        near_dots += 1
    if L == (image.getpixel((x, y + 1)) > G):
        near_dots += 1
    if L == (image.getpixel((x + 1, y - 1)) > G):
        near_dots += 1
    if L == (image.getpixel((x + 1, y)) > G):
        near_dots += 1
    if L == (image.getpixel((x + 1, y + 1)) > G):
        near_dots += 1

    if near_dots < N:
        return image.getpixel((x, (y - 1)))
    else:
        return None


# 降噪次数
def clear_noise(image, G, N, Z):
    draw = ImageDraw.Draw(image)
    for i in range(0, Z):
        for x in range(1, image.size[0] - 1):
            for y in range(1, image.size[1] - 1):
                color = get_pixel(image, x, y, G, N)
                if color is not None:
                    draw.point((x, y), color)


# 图片预处理
def image_preprocessing(image):
    grey_im = image.convert('L')
    binary_im = grey_im.point(get_table(threshold=150), "1")
    clear_noise(binary_im, 50, 4, 10)
    binary_im.show()
    binary_im.save('./PP_code.png')


# 判断所爬取的网址是否为今天这个日期
def verify_date(url):
    pattern = re.compile(r"\d{4}", re.S)
    res = re.search(pattern, url).group(0)
    today = datetime.date.today().strftime("%m%d")
    return res == today


# 发送邮件
def send_email(file_path_list):
    sender = EMAIL_HOST_USER
    receivers = EMAIL_TO

    # 创建一个带附件的实例
    message = MIMEMultipart()
    message['From'] = sender  # 发送者
    message['To'] = ';'.join(receivers)  # 接收者
    subject = "PP与PE爬虫"
    message['Subject'] = Header(subject, 'utf-8')

    # 邮件正文内容
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if len(file_path_list) == 0:
        content = "PP和PE最新网址都还没出来。\n" + now
    elif len(file_path_list) == 2:
        content = "请查看附件。\n" + now
    elif PP_AIM_FILE in file_path_list:
        content = "PE最新网址还没出来，PP粒文件在附件中。\n" + now
    else:
        content = "PP最新网址还没出来，PE粒文件在附件中。\n" + now
    message.attach(MIMEText(content, 'plain', 'utf-8'))

    # 构造附
    if len(file_path_list) > 0:
        for f in file_path_list:
            att = MIMEText(open(f, 'rb').read(), 'base64', 'utf-8')
            att["Content-Type"] = 'application/octet-stream'
            att.add_header("Content-Disposition", "attachment", filename=("gbk", "", f))
            message.attach(att)
    try:
        smtp = smtplib.SMTP()
        smtp.connect(host=EMAIL_HOST, port=EMAIL_PORT)
        smtp.login(user=EMAIL_HOST_USER, password=EMAIL_HOST_PASSWORD)
        smtp.sendmail(sender, receivers, message.as_string())
        smtp.quit()
        print("邮件发送成功。")
    except smtplib.SMTPException as e:
        print(e)
        print("邮件发送失败。")


# 验证码识别
def code_verify(img_url, code_verify_url):
    urlretrieve(img_url, './code.png')
    image = Image.open('./code.png')
    content = pytesseract.image_to_string(image)
    res = requests.get(code_verify_url.format(code=content))
    return res


# 获取PP或PE的最新网址
def get_url(search_url, keywords):
    response = parsel.Selector(
        requests.post(search_url, {'pageNo': '1', 'keyword': keywords}).text)
    url = response.xpath("//ul[@class='contentList']/li[1]//a/@href").get()
    return url


# 生成中间文件
def gen_temp_file(item, filename):
    wb = Workbook()
    ws = wb['Sheet']
    ws.title = "PP粒"
    title = item['title']
    content = item['content']
    ws.append(t for t in title)
    row_index = 2
    for c in content:
        if len(c) == len(title):
            for i in range(1, len(c) + 1):
                ws.cell(row_index, i).value = c[i - 1]
        elif len(c) == len(title) - 1:
            for i in range(1, len(c) + 1):
                ws.cell(row_index, i + 1).value = c[i - 1]
        elif len(c) == len(title) - 2:
            for i in range(1, len(c) + 1):
                ws.cell(row_index, i + 2).value = c[i - 1]
        else:
            for i in range(1, len(c) + 1):
                ws.cell(row_index, i + 3).value = c[i - 1]
        row_index += 1
    wb.save(filename)
    df = pd.read_excel(filename, sheet_name=ws.title, index_col=False)
    os.remove(filename)
    return df


# 生成最终文件
def gen_aim_file(additional_content, data_frame, max_col, origin_excel, origin_excel_sheet, AIM_FILE, num_type, cursor_for_insert, conn, type_num):
    num_type_dict = {}
    for nt in num_type:
        num_type_dict[nt["specifications_name"]] = nt["category"]
    equip_dyn = parse_en(additional_content, data_frame['装置动态'])
    enterprise_type = {}
    error_number = set()
    for key in equip_dyn:
        flag = 0
        for t in equip_dyn[key]:
            if t.startswith("停车"):
                enterprise_type[key] = "停车"
                flag = -1
                continue
            type_name = num_type_dict.get(t, None)
            if type_name:
                flag = 1
                enterprise_type[key] = type_name
                continue
        if flag == 0:
            for edk in equip_dyn[key]:
                error_number.add(edk)
            enterprise_type[key] = "找不到对应类型"
    for en in error_number:
        cursor_for_insert.execute("INSERT INTO specification_comparison (`specifications_name`, `category`, `type`) values (%s, %s, %s)", [en, "找不到对应类型", type_num])
    conn.commit()

    for i in range(2, max_col):
        origin_excel_sheet.cell(4, i).value = enterprise_type.get(origin_excel_sheet.cell(1, i).value)
    origin_excel.save(AIM_FILE)


# 生成模板excel和excel_sheet
def gen_excel_and_sheet(name, today):
    middle_file_name = name + '_' + str(today) + '.xlsx'
    if name == "PP":
        origin_excel = openpyxl.load_workbook(PP_EXCEL_TEMPLATE)
        origin_excel_sheet = origin_excel['PP排产']
    else:
        origin_excel = openpyxl.load_workbook(PE_EXCEL_TEMPLATE)
        origin_excel_sheet = origin_excel['PE排产']
    origin_excel_sheet.cell(4, 1).value = today.strftime("%Y/%m/%d")

    return middle_file_name, origin_excel, origin_excel_sheet
